# 단어별 오답 여부를 추가하기 위하여 파일 불러오는 방식 변경
#
import time
from openpyxl import load_workbook


def word_tester(start, last):
    # 점수
    score = 0
    t_score = last - start + 1
    word_cnt = 1

    excel_test = load_workbook(filename='./English_words.xlsx')
    word_sheet = excel_test['words']

    # 변수 first = 시작 단어, last = 끝 단어
    while start <= last:

        # 단어 리스트가 있는 엑셀 파일 구조(column순): 1-번호, 2-단어, 3-뜻, 4-오답여부
        print(word_cnt, ".", word_sheet.cell(start, 3).value)
        answer = input(" >> 영어 단어: ")

        # 정답인 경우, 정답 여부에 0 입력
        if answer == word_sheet.cell(start, 2).value:
            wrong_answer = 0
            print("정답입니다!")
            score += 1

        # 오답인 경우, 오답 여부에 1 입력
        else:
            wrong_answer = 1
            print("오답, 정답은 " + word_sheet.cell(start, 2).value)

        # 다음으로 넘어가기 전

        print("\n")

        word_sheet.cell(start, 4).value = wrong_answer
        start += 1
        word_cnt += 1

    # 테스트 종료 시 테스트 일시 입력
    word_sheet.cell((start-1), 5).value = time.strftime("%Y-%m-%d %H:%M", time.localtime(time.time()))
    excel_test.save('./English_words.xlsx')
    print("당신의 점수는", "%.1f" % ((score * 100) / t_score) + "점 입니다.\n")


def integer_input(msg):
    while True:
        try:
            n = int(input(msg))
            break
        except ValueError:
            print("\n 1일부터 30일 중에서 날짜를 입력하세요.\n")
    return n


# 테스트 Day 선택하기
while True:
    day_no = integer_input("영어 단어 연습할 Day를 선택해 주세요(학습을 종료하려면 99를 입력해주세요): ")

    if day_no == 1:
        print("\n선택하신 날짜는: ", day_no, "일\n")
        s = 1
        l = 60
        word_tester(s, l)

    elif 1 < day_no <= 30:
        print("\n선택하신 날짜는: ", day_no, "일\n")
        s = 1
        s += 60 * (day_no - 1)
        l = s + 59
        word_tester(s, l)

    elif day_no == 99:
        print("\n 수고하셨습니다. \n 학습을 종료합니다.")
        break

    else:
        print("\n날짜를 잘못 선택했습니다. ", day_no)
